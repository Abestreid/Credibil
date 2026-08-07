from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any

from credibil.domain.sync.errors import ParseError

if TYPE_CHECKING:
    from collections.abc import Iterator

logger = logging.getLogger(__name__)

# Expected column positions in the "Company" sheet (0-indexed)
COLUMN_MAP = {
    "IDNO": 0,
    "registration_date": 1,
    "full_name": 2,
    "legal_form_code": 3,
    "address": 4,
    "CUATM": 5,
    "directors": 6,
    "founders": 7,
    "activities_licensed": 8,
    "activities_unlicensed": 9,
    "liquidation_date": 10,
}


class XLSXCompanyParser:
    """Streaming parser for the CKAN XLSX bulk download.

    Opens the workbook and yields rows one-by-one from the "Company" sheet
    without loading the entire file into memory.
    """

    def __init__(self, filepath: str, sheet_name: str = "Company") -> None:
        self._filepath = filepath
        self._sheet_name = sheet_name

    def parse_rows(self) -> Iterator[dict[str, Any]]:
        """Yield dicts keyed by normalized column names. Skips header rows.

        The CKAN XLSX may have 1 or more header/title rows before actual data.
        We skip any row where the IDNO column is not purely numeric (Moldovan
        IDNOs are 10-13 digit numbers).
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self._filepath, read_only=True, data_only=True)
        except Exception as e:
            raise ParseError(self._filepath, str(e)) from e

        if self._sheet_name not in wb.sheetnames:
            raise ParseError(
                self._filepath,
                f"Sheet '{self._sheet_name}' not found. Available: {wb.sheetnames}",
            )

        ws = wb[self._sheet_name]

        for row in ws.iter_rows(values_only=True):
            if not row or row[COLUMN_MAP["IDNO"]] is None:
                continue

            idno_raw = str(row[COLUMN_MAP["IDNO"]]).strip()
            if not idno_raw or not idno_raw.isdigit():
                continue

            yield {
                "idno": idno_raw,
                "registration_date": row[COLUMN_MAP["registration_date"]],
                "full_name": str(row[COLUMN_MAP["full_name"]]).strip()
                if row[COLUMN_MAP["full_name"]]
                else "",
                "legal_form_code": str(row[COLUMN_MAP["legal_form_code"]]).strip()
                if row[COLUMN_MAP["legal_form_code"]]
                else "",
                "address": str(row[COLUMN_MAP["address"]]).strip()
                if row[COLUMN_MAP["address"]]
                else "",
                "cuatm": str(row[COLUMN_MAP["CUATM"]]).strip() if row[COLUMN_MAP["CUATM"]] else "",
                "directors": str(row[COLUMN_MAP["directors"]]).strip()
                if row[COLUMN_MAP["directors"]]
                else "",
                "founders": str(row[COLUMN_MAP["founders"]]).strip()
                if row[COLUMN_MAP["founders"]]
                else "",
                "activities_licensed": str(row[COLUMN_MAP["activities_licensed"]]).strip()
                if row[COLUMN_MAP["activities_licensed"]]
                else "",
                "activities_unlicensed": str(row[COLUMN_MAP["activities_unlicensed"]]).strip()
                if row[COLUMN_MAP["activities_unlicensed"]]
                else "",
                "liquidation_date": row[COLUMN_MAP["liquidation_date"]],
            }

        wb.close()

    def count_rows(self) -> int:
        """Count data rows (excluding header)."""
        count = 0
        for _ in self.parse_rows():
            count += 1
        return count
