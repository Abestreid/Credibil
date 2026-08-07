"""XLSX export generators for company and person entities.

Produces multi-sheet Excel workbooks with:
- Protected sheets to prevent accidental modification
- Auto-filter on data tables
- Alternating row colours for readability
- Locale-aware section headers (RO/RU)
"""

from __future__ import annotations

from datetime import date as _date
from datetime import datetime
from typing import Any


def _safe(v: Any) -> str:
    return str(v) if v else ""


def _fmt_date(v: Any) -> str:
    if not v:
        return "—"
    if isinstance(v, (datetime, _date)):
        return v.isoformat()[:10]
    return str(v)[:10]


def _fmt_money(v: Any) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):,.2f} MDL"
    except (ValueError, TypeError):
        return str(v)


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


# ── Label translations (same keys as PDF) ──────────────────────────────────

_LABELS: dict[str, dict[str, str]] = {
    "ro": {
        "report_title": "RAPORT DE DUE DILIGENCE",
        "sheet_info": "Informații raport",
        "sheet_summary": "Rezumat",
        "sheet_persons": "Persoane conectate",
        "sheet_risk": "Riscuri",
        "sheet_court": "Cazuri judiciare",
        "sheet_tenders": "Licitații",
        "sheet_companies": "Companii conectate",
        "label_name": "Denumire",
        "label_idno": "IDNO",
        "label_status": "Statut",
        "label_legal_form": "Formă juridică",
        "label_caem": "CAEM",
        "label_registration": "Data înregistrării",
        "label_address": "Adresă",
        "label_postal_code": "Cod poștal",
        "label_category": "Categorie",
        "label_tax_debt": "Datorie fiscală",
        "label_idnp": "IDNP",
        "label_type": "Tip",
        "label_nationality": "Naționalitate",
        "label_roles": "Relație",
        "label_person_name": "Nume persoană",
        "label_company_name": "Denumire companie",
        "label_company_status": "Statut companie",
        "label_risk_category": "Categorie risc",
        "label_risk_level": "Nivel risc",
        "label_case_number": "Număr dosar",
        "label_court_name": "Instanță",
        "label_case_type": "Tip caz",
        "label_tender_title": "Titlu licitație",
        "label_tender_value": "Valoare",
        "label_tender_status": "Statut",
        "generated_at": "Generat la",
        "footer_report": "Raport generat de platforma Credibil",
    },
    "ru": {
        "report_title": "ОТЧЁТ О ДЬЮ-ДИЛИДЖЕНС",
        "sheet_info": "Информация об отчёте",
        "sheet_summary": "Сводка",
        "sheet_persons": "Связанные лица",
        "sheet_risk": "Риски",
        "sheet_court": "Судебные дела",
        "sheet_tenders": "Тендеры",
        "sheet_companies": "Связанные компании",
        "label_name": "Наименование",
        "label_idno": "ИДНО",
        "label_status": "Статус",
        "label_legal_form": "Юридическая форма",
        "label_caem": "КЭАМ",
        "label_registration": "Дата регистрации",
        "label_address": "Адрес",
        "label_postal_code": "Почтовый индекс",
        "label_category": "Категория",
        "label_tax_debt": "Налоговая задолженность",
        "label_idnp": "ИНПН",
        "label_type": "Тип",
        "label_nationality": "Гражданство",
        "label_roles": "Связь",
        "label_person_name": "ФИО",
        "label_company_name": "Название компании",
        "label_company_status": "Статус компании",
        "label_risk_category": "Категория риска",
        "label_risk_level": "Уровень риска",
        "label_case_number": "Номер дела",
        "label_court_name": "Суд",
        "label_case_type": "Тип дела",
        "label_tender_title": "Название тендера",
        "label_tender_value": "Сумма",
        "label_tender_status": "Статус",
        "generated_at": "Дата формирования",
        "footer_report": "Отчёт сформирован платформой Credibil",
    },
}


def _l(lang: str, key: str) -> str:
    labels = _LABELS.get(lang, _LABELS["ro"])
    return labels.get(key, key)


# ── Style constants ─────────────────────────────────────────────────────────

_HDR_FILL_SOLID = "solid"
_HDR_FG = "1A56DB"
_BANNER_FG = "0F172A"
_ALT_FG = "F8FAFC"
_INFO_FG = "EFF6FF"
_PROTECT_PWD = "credibil"


def _make_wb_styles():
    from openpyxl.styles import Border, Font, PatternFill, Side

    HDR_FILL = PatternFill(_HDR_FILL_SOLID, fgColor=_HDR_FG)
    BANNER_FILL = PatternFill(_HDR_FILL_SOLID, fgColor=_BANNER_FG)
    ALT_FILL = PatternFill(_HDR_FILL_SOLID, fgColor=_ALT_FG)
    INFO_FILL = PatternFill(_HDR_FILL_SOLID, fgColor=_INFO_FG)

    TITLE_FONT = Font(color="FFFFFF", bold=True, size=15)
    HDR_FONT = Font(color="FFFFFF", bold=True, size=9)
    LABEL_FONT = Font(bold=True, size=9, color="64748B")
    BODY_FONT = Font(size=9)
    META_FONT = Font(bold=True, size=10, color="1A56DB")
    SMALL_FONT = Font(size=8, color="64748B")

    SIDE = Side(border_style="thin", color="E2E8F0")
    BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

    return {
        "HDR_FILL": HDR_FILL,
        "BANNER_FILL": BANNER_FILL,
        "ALT_FILL": ALT_FILL,
        "INFO_FILL": INFO_FILL,
        "TITLE_FONT": TITLE_FONT,
        "HDR_FONT": HDR_FONT,
        "LABEL_FONT": LABEL_FONT,
        "BODY_FONT": BODY_FONT,
        "META_FONT": META_FONT,
        "SMALL_FONT": SMALL_FONT,
        "BORDER": BORDER,
    }


def _col_width(ws, col: int, w: float) -> None:
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col)].width = w


def _write_header(ws, headers: list, widths: list | None, S: dict) -> None:
    from openpyxl.styles import Alignment
    ws.append(headers)
    row = ws.max_row
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=row, column=i)
        c.fill = S["HDR_FILL"]
        c.font = S["HDR_FONT"]
        c.border = S["BORDER"]
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    if widths:
        for i, w in enumerate(widths, 1):
            _col_width(ws, i, w)


def _write_row(ws, values: list, alt: bool, S: dict) -> None:
    from openpyxl.styles import Alignment
    ws.append(values)
    row = ws.max_row
    for i in range(1, len(values) + 1):
        c = ws.cell(row=row, column=i)
        c.font = S["BODY_FONT"]
        c.border = S["BORDER"]
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if alt:
            c.fill = S["ALT_FILL"]


def _write_kv(ws, label: str, value: str, alt: bool, S: dict) -> None:
    from openpyxl.styles import Alignment
    ws.append([label, value])
    row = ws.max_row
    lc = ws.cell(row=row, column=1)
    vc = ws.cell(row=row, column=2)
    lc.font = S["LABEL_FONT"]
    lc.border = S["BORDER"]
    lc.alignment = Alignment(horizontal="left", vertical="top")
    vc.font = S["BODY_FONT"]
    vc.border = S["BORDER"]
    vc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if alt:
        lc.fill = S["ALT_FILL"]
        vc.fill = S["ALT_FILL"]


def _write_section(ws, title: str, S: dict) -> None:
    from openpyxl.styles import Alignment
    row = ws.max_row + 2
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = title.upper()
    ws[f"A{row}"].font = S["META_FONT"]
    ws[f"A{row}"].fill = S["INFO_FILL"]
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16


def _protect_sheet(ws) -> None:
    ws.protection.sheet = True
    ws.protection.password = _PROTECT_PWD
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


# ── Company XLSX ────────────────────────────────────────────────────────────

def build_company_xlsx(company: dict, persons: list[dict], dashboard: dict | None,
                       lang: str = "ro") -> bytes:
    """Return a company due-diligence XLSX workbook as raw bytes."""
    import openpyxl
    from openpyxl.styles import Alignment

    S = _make_wb_styles()
    now = _now_iso()
    company_name = company.get("name_ro") or company.get("name_ru") or "—"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.properties.title = f"Credibil — {company_name}"
    wb.properties.subject = f"Due Diligence | {company.get('idno', '')}"
    wb.properties.creator = "Credibil Platform"
    wb.properties.description = f"Generated: {now}"

    # ── Sheet 1: Report Information ─────────────────────────────────────
    ws1 = wb.create_sheet(_l(lang, "sheet_info"))
    _col_width(ws1, 1, 28)
    _col_width(ws1, 2, 55)

    ws1.merge_cells("A1:B1")
    ws1["A1"] = _l(lang, "report_title")
    ws1["A1"].font = S["TITLE_FONT"]
    ws1["A1"].fill = S["BANNER_FILL"]
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 32

    ws1.append([])
    _write_section(ws1, _l(lang, "sheet_summary"), S)

    kv_data = [
        (_l(lang, "label_name"), company_name),
        (_l(lang, "label_idno"), company.get("idno", "—")),
        (_l(lang, "label_status"), (company.get("status") or "—").upper()),
        (_l(lang, "label_legal_form"), company.get("legal_form") or "—"),
        (_l(lang, "label_caem"), company.get("caem_description") or company.get("caem") or "—"),
        (_l(lang, "label_registration"), _fmt_date(company.get("registration_date"))),
        (_l(lang, "label_address"), company.get("legal_address") or "—"),
        (_l(lang, "label_postal_code"), company.get("postal_code") or "—"),
        (_l(lang, "label_category"), company.get("business_category") or "—"),
        (_l(lang, "label_tax_debt"), _fmt_money(company.get("tax_debt"))),
        (_l(lang, "generated_at"), now),
    ]
    for i, (label, value) in enumerate(kv_data):
        _write_kv(ws1, label, value, alt=(i % 2 == 1), S=S)

    _protect_sheet(ws1)

    # ── Sheet 2: Persons ────────────────────────────────────────────────
    ws2 = wb.create_sheet(_l(lang, "sheet_persons"))
    _write_header(ws2, [
        _l(lang, "label_person_name"),
        _l(lang, "label_idnp"),
        _l(lang, "label_roles"),
    ], [35, 20, 40], S)

    for i, p in enumerate(persons):
        _write_row(ws2, [
            p.get("person_name") or "—",
            p.get("person_idnp") or "—",
            ", ".join(p.get("roles_in_current", [])) or "—",
        ], alt=(i % 2 == 1), S=S)

    if persons:
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
    _protect_sheet(ws2)

    # ── Dashboard sheets ────────────────────────────────────────────────
    if dashboard:
        # Risk indicators
        risk_indicators = dashboard.get("risk_indicators") or []
        if risk_indicators:
            ws_risk = wb.create_sheet(_l(lang, "sheet_risk"))
            _write_header(ws_risk, [
                _l(lang, "label_risk_category"),
                _l(lang, "label_risk_level"),
            ], [40, 30], S)
            for i, r in enumerate(risk_indicators):
                _write_row(ws_risk, [
                    r.get("category") or "—",
                    (r.get("level") or "—").upper(),
                ], alt=(i % 2 == 1), S=S)
            if risk_indicators:
                ws_risk.freeze_panes = "A2"
                ws_risk.auto_filter.ref = ws_risk.dimensions
            _protect_sheet(ws_risk)

        # Court cases
        court_stats = dashboard.get("court_statistics") or {}
        if court_stats:
            ws_court = wb.create_sheet(_l(lang, "sheet_court"))
            _col_width(ws_court, 1, 28)
            _col_width(ws_court, 2, 40)
            _write_section(ws_court, _l(lang, "section_court"), S)
            total = court_stats.get("total_cases")
            if total is not None:
                _write_kv(ws_court, "Total", str(total), alt=False, S=S)
            pending = court_stats.get("pending_cases")
            if pending is not None:
                _write_kv(ws_court, "Pending", str(pending), alt=True, S=S)
            _protect_sheet(ws_court)

        # Tenders
        tender_stats = dashboard.get("tender_statistics") or {}
        if tender_stats:
            ws_tender = wb.create_sheet(_l(lang, "sheet_tenders"))
            _col_width(ws_tender, 1, 28)
            _col_width(ws_tender, 2, 40)
            _write_section(ws_tender, _l(lang, "section_tenders"), S)
            total_t = tender_stats.get("total_tenders")
            if total_t is not None:
                _write_kv(ws_tender, "Total", str(total_t), alt=False, S=S)
            total_val = tender_stats.get("total_value")
            if total_val is not None:
                _write_kv(ws_tender, _l(lang, "label_tender_value"),
                          _fmt_money(total_val), alt=True, S=S)
            _protect_sheet(ws_tender)

    return _save_wb(wb)


def _save_wb(wb) -> bytes:
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Person XLSX ─────────────────────────────────────────────────────────────

def build_person_xlsx(person: dict, companies: list[dict], lang: str = "ro") -> bytes:
    """Return a person due-diligence XLSX workbook as raw bytes."""
    import openpyxl
    from openpyxl.styles import Alignment

    S = _make_wb_styles()
    now = _now_iso()
    full_name = person.get("full_name") or "—"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.properties.title = f"Credibil — {full_name}"
    wb.properties.subject = f"Due Diligence | IDNP: {person.get('idnp', '')}"
    wb.properties.creator = "Credibil Platform"
    wb.properties.description = f"Generated: {now}"

    # ── Sheet 1: Report Information ─────────────────────────────────────
    ws1 = wb.create_sheet(_l(lang, "sheet_info"))
    _col_width(ws1, 1, 28)
    _col_width(ws1, 2, 55)

    ws1.merge_cells("A1:B1")
    ws1["A1"] = _l(lang, "report_title")
    ws1["A1"].font = S["TITLE_FONT"]
    ws1["A1"].fill = S["BANNER_FILL"]
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 32

    ws1.append([])
    _write_section(ws1, _l(lang, "sheet_summary"), S)

    kv_data = [
        (_l(lang, "label_name"), full_name),
        (_l(lang, "label_idnp"), person.get("idnp") or "—"),
        (_l(lang, "label_type"), person.get("person_type") or "—"),
        (_l(lang, "label_nationality"), person.get("nationality") or "—"),
        (_l(lang, "generated_at"), now),
    ]
    for i, (label, value) in enumerate(kv_data):
        _write_kv(ws1, label, value, alt=(i % 2 == 1), S=S)

    _protect_sheet(ws1)

    # ── Sheet 2: Connected Companies ────────────────────────────────────
    ws2 = wb.create_sheet(_l(lang, "sheet_companies"))
    _write_header(ws2, [
        _l(lang, "label_company_name"),
        _l(lang, "label_idno"),
        _l(lang, "label_company_status"),
        _l(lang, "label_roles"),
    ], [35, 20, 20, 30], S)

    for i, c in enumerate(companies):
        _write_row(ws2, [
            c.get("company_name") or "—",
            c.get("company_idno") or "—",
            c.get("company_status") or "—",
            ", ".join(c.get("roles", [])) or "—",
        ], alt=(i % 2 == 1), S=S)

    if companies:
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions
    _protect_sheet(ws2)

    return _save_wb(wb)
